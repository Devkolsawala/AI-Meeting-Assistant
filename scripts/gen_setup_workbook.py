"""Generates the MeetCopilot Phase 1 manual-setup + testing workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "DCE6F1"
ZEBRA = "F2F6FC"
GREY = "808080"
AMBER = "FFF2CC"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def style_header(ws, row, ncols, fill=BLUE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 26


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name=FONT, bold=True, color="FFFFFF", size=16)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=2 - 1, value=subtitle)
    s.font = Font(name=FONT, italic=True, color="404040", size=10)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20


def write_rows(ws, header_row, rows, widths, ncols, done_col=None,
               group_col=None, wrap_cols=None):
    wrap_cols = wrap_cols or []
    style_header(ws, header_row, ncols)
    r = header_row + 1
    last_group = None
    for row in rows:
        is_group = group_col is not None and row[group_col - 1] != last_group
        for c in range(1, ncols + 1):
            val = row[c - 1] if c - 1 < len(row) else None
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(
                horizontal="left", vertical="top",
                wrap_text=(c in wrap_cols))
            cell.border = border
            if (r - header_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
            if group_col is not None and c == group_col and is_group:
                cell.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        last_group = row[group_col - 1] if group_col is not None else None
        r += 1
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if done_col is not None:
        dv = DataValidation(type="list", formula1='"Pending,Done,N/A"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{chr(64 + done_col)}{header_row + 1}:{chr(64 + done_col)}{r - 1}")
        for rr in range(header_row + 1, r):
            ws.cell(row=rr, column=done_col, value="Pending")
    return r


# ---------------------------------------------------------------- Start Here
ws = wb.active
ws.title = "Start Here"
title_block(ws, "MeetCopilot — Phase 1 Manual Setup & Testing",
            "Everything you must do by hand (accounts, keys, config) plus how to test each milestone.", 3)
ws.sheet_view.showGridLines = False

intro = [
    ("", "", ""),
    ("HOW TO USE THIS WORKBOOK", "", ""),
    ("Tab", "What it covers", ""),
    ("Manual Setup", "Step-by-step external setup: Supabase, Google OAuth, AWS Bedrock, Deepgram, build/run. Mark each 'Done'.", ""),
    ("Env Variables", "Every environment variable: which file it goes in, public vs secret, where to get it.", ""),
    ("Milestone Tests", "How to verify each milestone (M1-M6) works, with expected results.", ""),
    ("Acceptance", "The Phase 1 acceptance criteria and how each is proven.", ""),
    ("", "", ""),
    ("RECOMMENDED ORDER", "", ""),
    ("1", "Create the 4 accounts (Supabase, AWS w/ Bedrock, Google Cloud, Deepgram).", ""),
    ("2", "Do the 'Manual Setup' tab top to bottom.", ""),
    ("3", "Fill both env files from the 'Env Variables' tab.", ""),
    ("4", "Restore Electron + package/install the desktop app.", ""),
    ("5", "Start the API and Web, then run the 'Milestone Tests'.", ""),
    ("", "", ""),
    ("RUN COMMANDS (run from the repo root: D:\\Dev\\AI-Meeting-Assistant)", "", ""),
    ("Purpose", "Command", "Notes"),
    ("Install deps", "pnpm install", "Restores Electron binary too (needed to run the desktop app)."),
    ("Typecheck / lint / build", "pnpm typecheck ; pnpm lint ; pnpm build", "All should pass with no errors."),
    ("Run API (backend)", "pnpm --filter @meetcopilot/api build ; pnpm --filter @meetcopilot/api start", "Listens on http://127.0.0.1:8787. Needs .env filled."),
    ("Run Web (login pages)", "pnpm --filter @meetcopilot/web dev", "Serves http://localhost:3000. Needs apps/web/.env.local."),
    ("Run Desktop (dev)", "pnpm --filter @meetcopilot/desktop build ; pnpm --filter @meetcopilot/desktop start", "For dev. Deep-link login is most reliable from the packaged build."),
    ("Package Desktop (installer)", '$env:ELECTRON_BUILDER_CACHE="D:\\eb-cache"; $env:TEMP="D:\\eb-tmp"; $env:TMP="D:\\eb-tmp"; pnpm --filter @meetcopilot/desktop package', "Produces apps/desktop/release/MeetCopilot Setup 0.0.0.exe. The env vars work around a cross-drive cache error on this PC."),
    ("", "", ""),
    ("TOOLCHAIN NOTES", "", ""),
    ("Node / pnpm", "Node 24.16.0, pnpm 11.5.1 (via corepack/Volta).", ""),
    ("If 'pnpm' is not found", "Prepend the working shim to PATH first: $env:Path=\"$env:USERPROFILE\\.pnpm-shim;$env:Path\"", "Only needed if the Volta pnpm shim fails in your shell."),
]
r = 4
for a, b, c in intro:
    ca = ws.cell(row=r, column=1, value=a)
    cb = ws.cell(row=r, column=2, value=b)
    cc = ws.cell(row=r, column=3, value=c)
    header_like = a in (
        "HOW TO USE THIS WORKBOOK", "RECOMMENDED ORDER",
        "RUN COMMANDS (run from the repo root: D:\\Dev\\AI-Meeting-Assistant)",
        "TOOLCHAIN NOTES")
    col_head = (a in ("Tab", "Purpose")) or (a == "1" and b.startswith("Create"))
    for cell in (ca, cb, cc):
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = Font(name=FONT, size=10)
    if header_like:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ca.font = Font(name=FONT, bold=True, size=12, color=NAVY)
        ca.fill = PatternFill("solid", fgColor=LIGHT)
        ws.row_dimensions[r].height = 22
    elif a in ("Tab", "Purpose"):
        for cell in (ca, cb, cc):
            cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLUE)
    if b and not header_like and a not in ("Tab", "Purpose"):
        cb.font = Font(name=FONT, size=10, color="C00000" if b.startswith("pnpm") or b.startswith("$env") else "000000")
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15, 30 if len(str(b)) > 70 else 18)
    r += 1
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 70
ws.column_dimensions["C"].width = 55


# ---------------------------------------------------------------- Manual Setup
ws = wb.create_sheet("Manual Setup")
ws.sheet_view.showGridLines = False
title_block(ws, "Manual Setup Tasks",
            "Do these in order. Mark the Status column as you go. 'Secret' values must never be committed.", 6)
hdr = 4
ws.cell(row=hdr, column=1, value="#")
ws.cell(row=hdr, column=2, value="Category")
ws.cell(row=hdr, column=3, value="Task")
ws.cell(row=hdr, column=4, value="Where / How")
ws.cell(row=hdr, column=5, value="Details & gotchas")
ws.cell(row=hdr, column=6, value="Status")

setup = [
    # Accounts
    (1, "Accounts", "Create a Supabase account & org", "https://supabase.com", "Free tier is fine for Phase 1."),
    (2, "Accounts", "Create/confirm an AWS account with Bedrock access", "https://console.aws.amazon.com", "Bedrock is region-specific; use us-east-1."),
    (3, "Accounts", "Create a Google Cloud account", "https://console.cloud.google.com", "Only needed for 'Sign in with Google'."),
    (4, "Accounts", "Create a Deepgram account", "https://console.deepgram.com", "Provides speech-to-text; needed for capture/transcripts."),
    # Supabase
    (5, "Supabase", "Create a new project", "Supabase dashboard > New project", "Pick region East US (us-east-1) to sit near the backend. Save the DB password."),
    (6, "Supabase", "Run the schema migration", "SQL Editor > paste supabase/migrations/0001_init_schema.sql > Run", "Creates users, personas, subscriptions, usage_sessions, usage_events + RLS. Additive & re-runnable."),
    (7, "Supabase", "Verify tables + RLS exist", "SQL Editor (see Milestone Tests > M2)", "All 5 tables should show rowsecurity = true."),
    (8, "Supabase", "Set Site URL", "Authentication > URL Configuration > Site URL", "Set to http://localhost:3000 for local dev."),
    (9, "Supabase", "Add redirect URL", "Authentication > URL Configuration > Redirect URLs", "Add http://localhost:3000/auth/callback (and your prod URL later). Do NOT add meetcopilot:// here."),
    (10, "Supabase", "Enable Email provider", "Authentication > Providers > Email", "Enables magic-link sign in. Dev mailer works out of the box; configure SMTP for real email."),
    (11, "Supabase", "Enable Google provider", "Authentication > Providers > Google", "Paste the Google Client ID + Secret from steps 15-16."),
    (12, "Supabase", "Copy Project URL & anon key", "Project Settings > Data API", "-> SUPABASE_URL, SUPABASE_ANON_KEY, NEXT_PUBLIC_* (PUBLIC, safe in clients)."),
    (13, "Supabase", "Copy service_role key", "Project Settings > API Keys", "-> SUPABASE_SERVICE_ROLE_KEY (SECRET, server only)."),
    (14, "Supabase", "Copy JWT secret", "Project Settings > JWT Keys (legacy JWT secret)", "-> SUPABASE_JWT_SECRET (SECRET). Backend uses it to verify user tokens."),
    # Google OAuth
    (15, "Google OAuth", "Configure OAuth consent screen", "Google Cloud > APIs & Services > OAuth consent screen", "External; add your email as a test user."),
    (16, "Google OAuth", "Create OAuth Client (Web application)", "Google Cloud > Credentials > Create credentials > OAuth client ID", "Authorized redirect URI: https://<project-ref>.supabase.co/auth/v1/callback . Copy Client ID + Secret into Supabase (step 11)."),
    # AWS Bedrock
    (17, "AWS Bedrock", "Use region us-east-1", "AWS console region switcher", "Model IDs/access are per-region; keep everything in us-east-1."),
    (18, "AWS Bedrock", "Enable model access", "Bedrock console > Model access", "Request/enable Amazon Nova Lite (fast lane) and Claude 3.5 Sonnet (smart lane)."),
    (19, "AWS Bedrock", "Create IAM credentials", "IAM > Users > create user with programmatic access", "Attach a policy allowing bedrock:InvokeModel and bedrock:InvokeModelWithResponseStream (Converse)."),
    (20, "AWS Bedrock", "Copy access keys", "IAM user > Security credentials", "-> AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY (SECRET). Set AWS_REGION=us-east-1."),
    # Deepgram
    (21, "Deepgram", "Create an API key", "Deepgram console > API Keys", "-> DEEPGRAM_API_KEY (SECRET, server only). Used by POST /stt-token."),
    # Env + build
    (22, "Env files", "Fill the root .env", "Copy .env.example -> .env, fill values", "See the 'Env Variables' tab. Never commit .env (it is gitignored)."),
    (23, "Env files", "Fill apps/web/.env.local", "Copy apps/web/.env.example -> apps/web/.env.local", "Holds NEXT_PUBLIC_SUPABASE_URL and NEXT_PUBLIC_SUPABASE_ANON_KEY."),
    (24, "Build & run", "Restore Electron binary", "pnpm install (from repo root)", "Earlier CI-style installs skipped it; this re-downloads it so the desktop app can launch."),
    (25, "Build & run", "Package + install the desktop app", "See 'Start Here' > Package Desktop, then run the generated Setup.exe", "Installing registers the meetcopilot:// scheme so deep-link login works reliably."),
]
write_rows(ws, hdr, setup,
           widths=[5, 14, 34, 40, 60, 11], ncols=6,
           done_col=6, group_col=2, wrap_cols=[3, 4, 5])


# ---------------------------------------------------------------- Env Variables
ws = wb.create_sheet("Env Variables")
ws.sheet_view.showGridLines = False
title_block(ws, "Environment Variables",
            "PUBLIC values are safe in clients. SECRET values are server-only and must never be committed or shipped to the desktop/web client.", 5)
hdr = 4
for i, h in enumerate(["Variable", "Goes in file", "Public / Secret", "Where to get it", "Example / Default"], start=1):
    ws.cell(row=hdr, column=i, value=h)

envs = [
    ("SUPABASE_URL", ".env (root)", "Public", "Supabase > Project Settings > Data API", "https://abcdxyz.supabase.co"),
    ("SUPABASE_ANON_KEY", ".env (root)", "Public", "Supabase > Project Settings > Data API", "eyJhbGci... (publishable)"),
    ("WEB_URL", ".env (root)", "Public", "Your web app URL", "http://localhost:3000"),
    ("API_URL", ".env (root)", "Public", "Your backend URL", "http://127.0.0.1:8787"),
    ("SUPABASE_SERVICE_ROLE_KEY", ".env (root)", "SECRET", "Supabase > Project Settings > API Keys", "(service_role secret)"),
    ("SUPABASE_JWT_SECRET", ".env (root)", "SECRET", "Supabase > Project Settings > JWT Keys", "(legacy JWT secret)"),
    ("AWS_REGION", ".env (root)", "Public", "Your AWS region", "us-east-1"),
    ("AWS_ACCESS_KEY_ID", ".env (root)", "SECRET", "AWS IAM user credentials", "AKIA..."),
    ("AWS_SECRET_ACCESS_KEY", ".env (root)", "SECRET", "AWS IAM user credentials", "(secret access key)"),
    ("BEDROCK_FAST_MODEL_ID", ".env (root)", "Public (optional)", "Bedrock model id (fast lane)", "amazon.nova-lite-v1:0 (default in code)"),
    ("BEDROCK_SMART_MODEL_ID", ".env (root)", "Public (optional)", "Bedrock model id (smart lane)", "anthropic.claude-3-5-sonnet-20240620-v1:0"),
    ("DEEPGRAM_API_KEY", ".env (root)", "SECRET", "Deepgram console > API Keys", "(deepgram key)"),
    ("ELEVENLABS_API_KEY", ".env (root)", "SECRET (optional)", "ElevenLabs (only if using that STT)", "(optional)"),
    ("SARVAM_API_KEY", ".env (root)", "SECRET (optional)", "Sarvam (only if using that STT)", "(optional)"),
    ("NEXT_PUBLIC_SUPABASE_URL", "apps/web/.env.local", "Public", "Same as SUPABASE_URL", "https://abcdxyz.supabase.co"),
    ("NEXT_PUBLIC_SUPABASE_ANON_KEY", "apps/web/.env.local", "Public", "Same as SUPABASE_ANON_KEY", "eyJhbGci..."),
]
end = write_rows(ws, hdr, envs, widths=[34, 22, 18, 42, 46], ncols=5,
                 wrap_cols=[4, 5])
# Colour the Public/Secret column
for rr in range(hdr + 1, end):
    val = str(ws.cell(row=rr, column=3).value)
    if val.startswith("SECRET"):
        ws.cell(row=rr, column=3).font = Font(name=FONT, size=10, bold=True, color="C00000")
    else:
        ws.cell(row=rr, column=3).font = Font(name=FONT, size=10, color="2E7D32")


# ---------------------------------------------------------------- Milestone Tests
ws = wb.create_sheet("Milestone Tests")
ws.sheet_view.showGridLines = False
title_block(ws, "Milestone Tests (M1-M6)",
            "How to verify each milestone. Run the API + Web + Desktop as needed (see Start Here > Run Commands).", 6)
hdr = 4
for i, h in enumerate(["Milestone", "Proves", "Prerequisites", "Test steps", "Expected result", "Status"], start=1):
    ws.cell(row=hdr, column=i, value=h)

tests = [
    ("M1 — Monorepo + CI",
     "Workspace builds; CI runs on every push.",
     "GitHub repo connected.",
     "1) Locally: pnpm install, then pnpm typecheck, pnpm lint, pnpm build.\n"
     "2) Push to GitHub and open the Actions tab > 'CI / verify' run.",
     "All three local commands exit 0. The GitHub Actions run is green (install -> typecheck -> lint -> build).",
     "Pending"),
    ("M2 — Database schema + RLS",
     "All 5 tables exist with Row Level Security.",
     "Migration run in Supabase (Setup step 6).",
     "Run in Supabase SQL Editor:\n"
     "select tablename, rowsecurity from pg_tables where schemaname='public' order by tablename;\n"
     "select tablename, policyname, cmd from pg_policies where schemaname='public' order by tablename;",
     "users, personas, subscriptions, usage_sessions, usage_events all show rowsecurity = true, and policies are listed for each.",
     "Pending"),
    ("M3 — Auth (deep-link PKCE)",
     "Browser login returns to the app; token stored encrypted.",
     "Supabase auth configured (Google+email, redirect URL); env filled; desktop app INSTALLED via Setup.exe; Web running.",
     "1) Launch MeetCopilot. 2) Click 'Sign in'. 3) In the browser, sign in with Google or request an email magic link. 4) Approve returning to MeetCopilot.\n"
     "5) Check %APPDATA%\\MeetCopilot\\auth.enc exists. 6) Click 'Sign out'.",
     "Overlay shows 'Signed in as <email>'. auth.enc exists and is NOT human-readable (encrypted via safeStorage). Sign out removes it.",
     "Pending"),
    ("M4 — Backend /infer + /stt-token",
     "Authenticated streaming inference + STT token issuance.",
     "API running; .env has Supabase JWT + AWS Bedrock; Bedrock model access enabled.",
     "STT: curl -X POST http://127.0.0.1:8787/stt-token\n"
     "INFER: get a valid Supabase access token (e.g. from the signed-in app), then:\n"
     'curl -N -X POST http://127.0.0.1:8787/infer -H "Authorization: Bearer <token>" -H "Content-Type: application/json" -d "{\\"context\\":[{\\"speaker\\":\\"them\\",\\"text\\":\\"Tell me about yourself\\"}]}"',
     "/stt-token returns JSON with accessToken. /infer streams SSE 'data: {\"delta\":...}' lines ending with 'data: [DONE]'. Without a valid token /infer returns 401.",
     "Pending"),
    ("M5 — Desktop -> backend (end to end)",
     "Hotkey sends context to /infer; overlay streams the answer.",
     "API + Web running; signed in; capture started.",
     "1) Click 'Start capture' and grant mic + screen-with-audio. 2) Speak (You) and/or play meeting audio (Them) so transcript lines appear.\n"
     "3) Press Ctrl+Enter (or click 'Ask'). ",
     "The Answer panel streams a useful response, beginning in under ~2 seconds. (Deepgram connects via POST /stt-token.)",
     "Pending"),
    ("M6 — Interview persona",
     "Persona system prompt + notes template injected server-side.",
     "Same as M5 (desktop sends persona 'interview' by default).",
     "During a mock interview, let 'Them' ask a question (e.g. 'Tell me about a challenge you faced'), then press Ctrl+Enter.\n"
     "Optional backend check: send /infer with \"persona\":\"interview\" vs without and compare tone.",
     "The answer is concise, first-person, and STAR-structured (Situation/Task/Action/Result) — i.e. interview-tuned, not generic.",
     "Pending"),
]
end = write_rows(ws, hdr, tests, widths=[24, 30, 30, 62, 50, 11], ncols=6,
                 done_col=6, wrap_cols=[2, 3, 4, 5])
for rr in range(hdr + 1, end):
    ws.cell(row=rr, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.row_dimensions[rr].height = 96


# ---------------------------------------------------------------- Acceptance
ws = wb.create_sheet("Acceptance")
ws.sheet_view.showGridLines = False
title_block(ws, "Phase 1 Acceptance Criteria",
            "The bar for 'Phase 1 done'. Each maps to a milestone test.", 4)
hdr = 4
for i, h in enumerate(["Criterion", "How it is proven", "Related test", "Status"], start=1):
    ws.cell(row=hdr, column=i, value=h)
acc = [
    ("Real mock call -> useful streamed answer in under 2s", "Run a mock meeting, press Ctrl+Enter, watch the Answer panel.", "M5 / M6", "Pending"),
    ("Login works", "Browser sign-in returns to the app; status shows the email.", "M3", "Pending"),
    ("No secrets in the client", "Only SUPABASE_URL + anon key reach desktop/web; provider & service-role keys stay server-side.", "M4 (design)", "Done"),
    ("CI green", "GitHub Actions 'CI / verify' passes on push.", "M1", "Pending"),
]
end = write_rows(ws, hdr, acc, widths=[46, 60, 16, 11], ncols=4,
                 done_col=4, wrap_cols=[1, 2])

OUT = r"D:\Dev\AI-Meeting-Assistant\docs\MeetCopilot-Phase1-Setup-and-Testing.xlsx"
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("Saved:", OUT)
